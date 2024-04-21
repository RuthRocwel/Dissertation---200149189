import openpyxl 
import matplotlib.pyplot as plt
import torch
import torch.nn as nn
import numpy as np
import pandas as pd


# Load the xlsm file
workbook = openpyxl.load_workbook('BirminghamCityTrafficCountData.xlsm', keep_vba=True)

# Access the sheets
sheet1 = workbook['ExtractedDataTotal1']
sheet2 = workbook['ExtractedDataTotal2']

# Initialize lists to store data for Sheet1
time1 = []
total1_sheet1 = []
total2_sheet1 = []

# Iterate over rows and store values from specific columns for Sheet1
for row in sheet1.iter_rows(min_row=2, values_only=True):  # Start from row 2
    time1.append(row[0].strftime('%H:%M'))   # Convert time to string in 'HH:MM' format
    total1_sheet1.append(row[1])
    total2_sheet1.append(row[2])

# Initialize lists to store data for Sheet2
time2 = []
total1_sheet2 = []
total2_sheet2 = []

# Iterate over rows and store values from specific columns for Sheet2
for row in sheet2.iter_rows(min_row=2, values_only=True):  # Start from row 2
    time2.append(row[0].strftime('%H:%M'))   # Convert time to string in 'HH:MM' format
    total1_sheet2.append(row[1])
    total2_sheet2.append(row[2])

# Convert data to PyTorch tensors
X_sheet1 = torch.tensor(total1_sheet1, dtype=torch.float32).view(-1, 1)  # Reshape to (N, 1) tensor
y_sheet1 = torch.tensor(total2_sheet1, dtype=torch.float32).view(-1, 1)  # Reshape to (N, 1) tensor

X_sheet2 = torch.tensor(total1_sheet2, dtype=torch.float32).view(-1, 1)  # Reshape to (N, 1) tensor
y_sheet2 = torch.tensor(total2_sheet2, dtype=torch.float32).view(-1, 1)  # Reshape to (N, 1) tensor

# Define the neural network model
class TrafficPredictor(nn.Module):
    def __init__(self):
        super(TrafficPredictor, self).__init__()
        self.fc1 = nn.Linear(1, 64)
        self.fc2 = nn.Linear(64, 1)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = self.fc2(x)
        return x

# Create instances of the model
model1 = TrafficPredictor()
model2 = TrafficPredictor()

# Define the loss function and optimizer
criterion = nn.MSELoss()
optimizer1 = torch.optim.Adam(model1.parameters(), lr=0.001)
optimizer2 = torch.optim.Adam(model2.parameters(), lr=0.001)

# Train the models
for epoch in range(100):
    # Train model for Sheet1
    optimizer1.zero_grad()
    outputs1 = model1(X_sheet1)
    loss1 = criterion(outputs1, y_sheet1)
    loss1.backward()
    optimizer1.step()

    # Train model for Sheet2
    optimizer2.zero_grad()
    outputs2 = model2(X_sheet2)
    loss2 = criterion(outputs2, y_sheet2)
    loss2.backward()
    optimizer2.step()

    if (epoch+1) % 10 == 0:
        print(f'Epoch [{epoch+1}/100]')
        print(f'  Sheet1 Loss: {loss1.item():.4f}, Sheet2 Loss: {loss2.item():.4f}')

# Make predictions
predictions_sheet1 = model1(X_sheet1).detach().numpy()
predictions_sheet2 = model2(X_sheet2).detach().numpy()

print("predicted values 1:", predictions_sheet1)
print("predicted values 2:", predictions_sheet2)


# Plot and print predictions for both sheets
plt.figure(figsize=(15, 5))

# Plot and print predictions for Sheet1
plt.subplot(1, 2, 1)
plt.scatter(total1_sheet1, total2_sheet1, label='Actual')
plt.plot(total1_sheet1, predictions_sheet1, color='red', label='Predicted')
plt.xlabel('Total1 (Sheet1)')
plt.ylabel('Total2 (Sheet1)')
plt.title('Predicted vs Actual (Sheet1)')
plt.legend()

# Plot and print predictions for Sheet2
plt.subplot(1, 2, 2)
plt.scatter(total1_sheet2, total2_sheet2, label='Actual')
plt.plot(total1_sheet2, predictions_sheet2, color='red', label='Predicted')
plt.xlabel('Total1 (Sheet2)')
plt.ylabel('Total2 (Sheet2)')
plt.title('Predicted vs Actual (Sheet2)')
plt.legend()

plt.tight_layout()
plt.show()




# Function to calculate MAE
def mean_absolute_error(actual_values, predicted_values):

    # Ensure the lengths of actual and predicted values are the same
    if len(actual_values) != len(predicted_values):
        raise ValueError("Lengths of actual and predicted values must be the same.")
    
    # Calculate the absolute differences between actual and predicted values
    absolute_errors = [abs(actual - predicted) for actual, predicted in zip(actual_values, predicted_values)]
    
    # Calculate the mean of absolute errors
    mae = sum(absolute_errors) / len(absolute_errors)
    
    return mae

# Calculate MAE using actual and predicted values
mae_sheet1 = mean_absolute_error(X_sheet1, predictions_sheet1)
mae_sheet2 = mean_absolute_error(X_sheet2, predictions_sheet2)

print("Mean Absolute Error (MAE) for Sheet 1:", mae_sheet1)
print("Mean Absolute Error (MAE) for Sheet 2:", mae_sheet2)
