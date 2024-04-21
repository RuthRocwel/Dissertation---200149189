import tkinter as tk
from tkinter.scrolledtext import ScrolledText
from tkinter import ttk
import openpyxl 
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import torch
import torch.nn as nn
import numpy as np
import pandas as pd

class TrainingGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Training Progress")
        self.root.geometry("800x600")

        self.tabControl = ttk.Notebook(self.root)
        self.tabControl.pack(expand=1, fill="both")

        self.tab1 = ttk.Frame(self.tabControl)
        self.tabControl.add(self.tab1, text="Sheet 1")

        self.tab2 = ttk.Frame(self.tabControl)
        self.tabControl.add(self.tab2, text="Sheet 2")

        self.text_area1 = ScrolledText(self.tab1, wrap=tk.WORD, width=80, height=30)
        self.text_area1.pack(pady=20)

        self.text_area2 = ScrolledText(self.tab2, wrap=tk.WORD, width=80, height=30)
        self.text_area2.pack(pady=20)

        self.train_model_sheet1()
        self.train_model_sheet2()

    def print_to_gui(self, message, tab):
        if tab == 1:
            text_area = self.text_area1
        elif tab == 2:
            text_area = self.text_area2
        else:
            raise ValueError("Invalid tab number. Must be 1 or 2.")
        
        text_area.insert(tk.END, message)
        text_area.see(tk.END)

    def train_model_sheet1(self):
        # Load the xlsm file
        workbook = openpyxl.load_workbook('BirminghamCityTrafficCountData.xlsm', keep_vba=True)

        # Access the sheet
        sheet = workbook['ExtractedDataTotal1']

        # Initialize lists to store data
        time = []
        total1 = []
        total2 = []

        # Iterate over rows and store values
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2
            time.append(row[0].strftime('%H:%M'))   # Convert time to string in 'HH:MM' format
            total1.append(row[1])
            total2.append(row[2])

        # Convert data to PyTorch tensors
        X = torch.tensor(total1, dtype=torch.float32).view(-1, 1)  # Reshape to (N, 1) tensor
        y = torch.tensor(total2, dtype=torch.float32).view(-1, 1)  # Reshape to (N, 1) tensor

        # Define the neural network model
        class TrafficPredictor(nn.Module):
            def __init__(self):
                super(TrafficPredictor, self).__init__()
                self.fc1 = nn.Linear(1, 64)
                self.fc2 = nn.Linear(64, 1)

            def forward(self, x):
                x = torch.relu(self.fc1(x))
                x = self.fc2(x)
                return x

        # Create instance of the model
        model = TrafficPredictor()

        # Define the loss function and optimizer
        criterion = nn.MSELoss()
        optimizer = torch.optim.Adam(model.parameters(), lr=0.001)

        # Train the model
        for epoch in range(100):
            # Train model
            optimizer.zero_grad()
            outputs = model(X)
            loss = criterion(outputs, y)
            loss.backward()
            optimizer.step()

            if (epoch+1) % 10 == 0:
                progress_message = f'Epoch [{epoch+1}/100]\n'
                progress_message += f'  Loss: {loss.item():.4f}\n'
                self.print_to_gui(progress_message, tab=1)

        # Make predictions
        predictions = model(X).detach().numpy()

        result_message = "Predictions:\n"
        result_message += "Predicted values: \n" + str(predictions) + "\n\n"
        self.print_to_gui(result_message, tab=1)

        # Plot and print predictions
        plt.figure(figsize=(8, 6))
        plt.scatter(total1, total2, label='Actual')
        plt.plot(total1, predictions, color='red', label='Predicted')
        plt.xlabel('Total1')
        plt.ylabel('Total2')
        plt.title('Predicted vs Actual')
        plt.legend()
        plt.tight_layout()

        canvas = FigureCanvasTkAgg(plt.gcf(), master=self.tab1)
        canvas.draw()
        canvas.get_tk_widget().pack()

    def train_model_sheet2(self):
        # Load the xlsm file
        workbook = openpyxl.load_workbook('BirminghamCityTrafficCountData.xlsm', keep_vba=True)

        # Access the sheet
        sheet = workbook['ExtractedDataTotal2']

        # Initialize lists to store data
        time = []
        total1 = []
        total2 = []

        # Iterate over rows and store values
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2
            time.append(row[0].strftime('%H:%M'))   # Convert time to string in 'HH:MM' format
            total1.append(row[1])
            total2.append(row[2])

        # Convert data to PyTorch tensors
        X = torch.tensor(total1, dtype=torch.float32).view(-1, 1)  # Reshape to (N, 1) tensor
        y = torch.tensor(total2, dtype=torch.float32).view(-1, 1)  # Reshape to (N, 1) tensor

        # Define the neural network model
        class TrafficPredictor(nn.Module):
            def __init__(self):
                super(TrafficPredictor, self).__init__()
                self.fc1 = nn.Linear(1, 64)
                self.fc2 = nn.Linear(64, 1)

            def forward(self, x):
                x = torch.relu(self.fc1(x))
                x = self.fc2(x)
                return x

        # Create instance of the model
        model = TrafficPredictor()

        # Define the loss function and optimizer
        criterion = nn.MSELoss()
        optimizer = torch.optim.Adam(model.parameters(), lr=0.001)

        # Train the model
        for epoch in range(100):
            # Train model
            optimizer.zero_grad()
            outputs = model(X)
            loss = criterion(outputs, y)
            loss.backward()
            optimizer.step()

            if (epoch+1) % 10 == 0:
                progress_message = f'Epoch [{epoch+1}/100]\n'
                progress_message += f'  Loss: {loss.item():.4f}\n'
                self.print_to_gui(progress_message, tab=2)

        # Make predictions
        predictions = model(X).detach().numpy()

        result_message = "Predictions:\n"
        result_message += "Predicted values: \n" + str(predictions) + "\n\n"
        self.print_to_gui(result_message, tab=2)

        # Plot and print predictions
        plt.figure(figsize=(8, 6))
        plt.scatter(total1, total2, label='Actual')
        plt.plot(total1, predictions, color='red', label='Predicted')
        plt.xlabel('Total1')
        plt.ylabel('Total2')
        plt.title('Predicted vs Actual')
        plt.legend()
        plt.tight_layout()

        canvas = FigureCanvasTkAgg(plt.gcf(), master=self.tab2)
        canvas.draw()
        canvas.get_tk_widget().pack()

if __name__ == "__main__":
    root = tk.Tk()
    app = TrainingGUI(root)
    root.mainloop()
